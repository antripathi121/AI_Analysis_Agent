import os
import re
from io import BytesIO
from datetime import timedelta

import pandas as pd
import streamlit as st

from rollout_rules import (
    build_current_coverage_segments,
    build_p1_coverage_segments,
    build_p2_coverage_segments,
    store_is_allowed,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Retail Campaign Intelligence Engine", page_icon="📊", layout="wide")

LIBRARY_FILE = "campaign_library.xlsx"

PHASE_1_STORES = ["Hyrum"]

PHASE_2_STORES = [
    "Hyrum",
    "Fillmore",
    "Nephi",
    "Gunnison",
    "Mt Pleasant",
]

CURRENT_PHASES = [
    {
        "name": "Phase 1",
        "start": pd.Timestamp("2026-01-01"),
        "end": pd.Timestamp("2026-04-01"),
        "stores": PHASE_1_STORES,
    },
    {
        "name": "Phase 2",
        "start": pd.Timestamp("2026-04-02"),
        "end": pd.Timestamp("2026-06-15"),
        "stores": PHASE_2_STORES,
    },
    {
        "name": "Phase 3",
        "start": pd.Timestamp("2026-06-16"),
        "end": pd.Timestamp("2099-12-31"),
        "stores": "ALL",
    },
]

P1_PHASES = [
    {
        "name": "P1 Phase 1",
        "start": pd.Timestamp("2025-11-16"),
        "end": pd.Timestamp("2026-04-01"),
        "stores": PHASE_1_STORES,
    },
    {
        "name": "P1 Phase 2",
        "start": pd.Timestamp("2026-04-02"),
        "end": pd.Timestamp("2026-06-14"),
        "stores": PHASE_2_STORES,
    },
    {
        "name": "P1 Phase 3",
        "start": pd.Timestamp("2026-06-15"),
        "end": pd.Timestamp("2099-12-31"),
        "stores": "ALL",
    },
]


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_cpg_name(name):
    text = clean_text(name).lower()
    text = text.replace("_salesreport", "")
    text = text.replace("salesreport", "")
    text = text.replace("_sales report", "")
    text = text.replace("sales report", "")

    remove_words = [
        "the", "company", "co", "inc", "incorporated",
        "llc", "ltd", "limited", "corp", "corporation"
    ]

    for char in [",", ".", "-", "_", "(", ")", "&"]:
        text = text.replace(char, " ")

    words = text.split()
    words = [w for w in words if w not in remove_words]

    return " ".join(words).strip()


def normalize_store_text(value):
    text = clean_text(value).lower()
    for char in [",", ".", "-", "_", "(", ")", "&"]:
        text = text.replace(char, " ")
    return " ".join(text.split())


def excel_date(dt):
    dt = pd.to_datetime(dt)
    return f"{dt.month}/{dt.day}/{str(dt.year)[-2:]}"


def pct_change(current, previous):
    if previous == 0 or pd.isna(previous):
        return None
    return ((current - previous) / previous) * 100


def format_pct(value):
    if value is None or pd.isna(value):
        return "N/A"
    arrow = "▲" if value >= 0 else "▼"
    return f"{arrow} {value:,.1f}%"


def find_header_row(raw_df):
    for i in range(min(25, len(raw_df))):
        row_values = [clean_text(v).lower() for v in raw_df.iloc[i].tolist()]

        has_offer = any(v == "offer name" for v in row_values)
        has_cpg = any(v in ["cpg", "cpg name"] for v in row_values)
        has_start = any(v == "start date" for v in row_values)
        has_end = any(v == "end date" for v in row_values)

        if has_offer and has_cpg and has_start and has_end:
            return i

    raise Exception("Could not find redemption header row. Expected Offer Name, CPG/CPG Name, Start Date, End Date.")


def get_column(df, possible_names, default_value=0):
    normalized_cols = {clean_text(c).lower(): c for c in df.columns}

    for name in possible_names:
        key = name.lower()
        if key in normalized_cols:
            return df[normalized_cols[key]]

    return pd.Series([default_value] * len(df))


def load_redemption_report(uploaded_file):
    raw = pd.read_excel(uploaded_file, header=None)
    header_row = find_header_row(raw)

    df = pd.read_excel(uploaded_file, header=header_row)
    df.columns = df.columns.astype(str).str.strip()

    cleaned = pd.DataFrame()

    cleaned["Offer Name"] = get_column(df, ["Offer Name"])
    cleaned["CPG"] = get_column(df, ["CPG", "CPG Name"])
    cleaned["Type"] = get_column(df, ["Type"])
    cleaned["CB Value"] = get_column(df, ["CB Value"])
    cleaned["Budget"] = get_column(df, ["Budget"])
    cleaned["Start Date"] = get_column(df, ["Start Date"])
    cleaned["End Date"] = get_column(df, ["End Date"])
    cleaned["Avail Budget"] = get_column(df, ["Avail Budget", "Available Budget"])
    cleaned["Activations"] = get_column(df, ["Activations"])
    cleaned["Redemptions"] = get_column(df, ["Redemptions"])
    cleaned["CB Amount"] = get_column(df, ["CB Amount"])
    cleaned["Total"] = get_column(df, ["Total"])
    cleaned["Unique"] = get_column(df, ["Unique"])
    cleaned["New"] = get_column(df, ["New"])

    cleaned = cleaned[cleaned["CPG"].notna()]
    cleaned = cleaned[cleaned["CPG"].astype(str).str.strip() != ""]
    cleaned = cleaned[cleaned["CPG"].astype(str).str.strip() != "0"]

    cleaned["Start Date"] = pd.to_datetime(cleaned["Start Date"], errors="coerce")
    cleaned["End Date"] = pd.to_datetime(cleaned["End Date"], errors="coerce")

    cleaned = cleaned[cleaned["Start Date"].notna()]
    cleaned = cleaned[cleaned["End Date"].notna()]

    cleaned.to_excel(LIBRARY_FILE, index=False)
    return cleaned


def load_campaign_library():
    if not os.path.exists(LIBRARY_FILE):
        raise Exception("Please upload Daily Redemption Report first to create campaign library.")

    df = pd.read_excel(LIBRARY_FILE)
    df["Start Date"] = pd.to_datetime(df["Start Date"])
    df["End Date"] = pd.to_datetime(df["End Date"])
    return df


def read_sales_sheet(uploaded_file):
    return pd.read_excel(uploaded_file, sheet_name="All Brands Weekly", header=None)


def get_sales_cpg(sheet):
    value = clean_text(sheet.iloc[0, 0])
    if value.lower().startswith("cpg name"):
        value = value.split(":", 1)[1].strip()

    value = re.sub(r"_?sales\s*report$", "", value, flags=re.IGNORECASE).strip()
    value = value.replace("_SalesReport", "").replace("_salesreport", "").strip()

    return value


def find_campaign(library_df, sales_cpg):
    sales_cpg_clean = clean_text(sales_cpg).lower()

    exact = library_df[
        library_df["CPG"].astype(str).str.strip().str.lower() == sales_cpg_clean
    ]

    if not exact.empty:
        return exact.iloc[0], "Exact CPG Match"

    sales_cpg_normalized = normalize_cpg_name(sales_cpg)

    library_df = library_df.copy()
    library_df["Normalized CPG"] = library_df["CPG"].apply(normalize_cpg_name)

    normalized_match = library_df[
        library_df["Normalized CPG"] == sales_cpg_normalized
    ]

    if not normalized_match.empty:
        return normalized_match.iloc[0], "Normalized CPG Match"

    possible_cpgs = library_df["CPG"].dropna().astype(str).head(15).tolist()

    raise Exception(
        f"No matching campaign found for Sales Report CPG: {sales_cpg}. "
        f"Normalized as: {sales_cpg_normalized}. "
        f"Available CPG examples in library: {possible_cpgs}"
    )


def find_row_containing(sheet, keyword):
    keyword = keyword.lower()

    for row in range(sheet.shape[0]):
        for col in range(sheet.shape[1]):
            value = clean_text(sheet.iat[row, col]).lower()
            if keyword in value:
                return row, col

    raise Exception(f"Could not find section: {keyword}")


def find_store_id_row(sheet, start_row):
    for row in range(start_row, min(start_row + 20, sheet.shape[0])):
        for col in range(sheet.shape[1]):
            value = clean_text(sheet.iat[row, col]).lower().replace(" ", "")
            if value == "storeid":
                return row, col

    raise Exception("Could not find Store ID row.")


def get_week_columns(sheet, header_row):
    week_cols = {}

    for col in range(1, sheet.shape[1]):
        value = sheet.iat[header_row, col]
        date_value = pd.to_datetime(value, errors="coerce")

        if not pd.isna(date_value):
            week_cols[col] = date_value.normalize()

    if not week_cols:
        raise Exception("No weekly date columns found.")

    return week_cols


def get_data_end_row(sheet, start_row, store_col):
    row = start_row

    while row < sheet.shape[0]:
        store_value = sheet.iat[row, store_col]
        if pd.isna(store_value) or str(store_value).strip() == "":
            break
        row += 1

    return row


def store_is_allowed(store_value, active_stores):
    if active_stores == "ALL":
        return True

    store_text = normalize_store_text(store_value)

    for store in active_stores:
        store_name = normalize_store_text(store)
        if store_name in store_text:
            return True

    return False


def build_segments_from_timeline(period_start, period_end, timeline, label):
    period_start = pd.to_datetime(period_start).normalize()
    period_end = pd.to_datetime(period_end).normalize()

    segments = []

    for phase in timeline:
        overlap_start = max(period_start, phase["start"])
        overlap_end = min(period_end, phase["end"])

        if overlap_start <= overlap_end:
            segments.append({
                "period": label,
                "phase": phase["name"],
                "start": overlap_start,
                "end": overlap_end,
                "stores": phase["stores"]
            })

    if not segments:
        raise Exception(f"{label} does not overlap with configured store rollout rules.")

    return segments


def build_current_coverage_segments(current_start, current_end):
    return build_segments_from_timeline(
        current_start,
        current_end,
        CURRENT_PHASES,
        "Current"
    )


def build_p1_coverage_segments(p1_start, p1_end):
    return build_segments_from_timeline(
        p1_start,
        p1_end,
        P1_PHASES,
        "P1"
    )


def build_p2_coverage_segments(p2_start, p2_end):
    p2_start = pd.to_datetime(p2_start).normalize()
    p2_end = pd.to_datetime(p2_end).normalize()

    return [{
        "period": "P2",
        "phase": "P2 Hyrum Historical Rule",
        "start": p2_start,
        "end": p2_end,
        "stores": PHASE_1_STORES
    }]


def selected_columns_for_period(week_cols, start_date, end_date):
    """
    Current Period weekly mapping rule.

    Weekly sales report dates are week-start dates.

    Start Date:
    Map to the latest week start on or before the offer start date.

    End Date:
    Map to the latest week start on or before the offer end date.
    If the offer end date falls exactly on a week start date,
    use the previous week's start date to avoid adding a full extra week.

    This function is mainly used to determine the Current Period weeks.
    """

    start_date = pd.to_datetime(start_date).normalize()
    end_date = pd.to_datetime(end_date).normalize()

    sorted_weeks = sorted(
        [(col, pd.to_datetime(week_start).normalize()) for col, week_start in week_cols.items()],
        key=lambda x: x[1]
    )

    start_candidates = [
        (col, week_start)
        for col, week_start in sorted_weeks
        if week_start <= start_date
    ]

    if not start_candidates:
        return []

    mapped_start_week = start_candidates[-1][1]

    all_week_starts = [week_start for _, week_start in sorted_weeks]

    if end_date in all_week_starts:
        end_candidates = [
            (col, week_start)
            for col, week_start in sorted_weeks
            if week_start < end_date
        ]
    else:
        end_candidates = [
            (col, week_start)
            for col, week_start in sorted_weeks
            if week_start <= end_date
        ]

    if not end_candidates:
        return []

    mapped_end_week = end_candidates[-1][1]

    selected_cols = [
        col
        for col, week_start in sorted_weeks
        if mapped_start_week <= week_start <= mapped_end_week
    ]

    return selected_cols


def selected_columns_before_current(week_cols, current_start_date, current_week_count):
    """
    P1 rule.

    P1 must use the exact same number of counted weekly buckets
    as the Current Period.

    It selects the immediately preceding N weekly buckets before
    the first Current Period week.

    Example:
    Current = 12 weeks starting 2/15/26
    P1 = previous 12 weeks ending 2/8/26
    """

    current_start_date = pd.to_datetime(current_start_date).normalize()

    sorted_weeks = sorted(
        [(col, pd.to_datetime(week_start).normalize()) for col, week_start in week_cols.items()],
        key=lambda x: x[1]
    )

    current_start_candidates = [
        (col, week_start)
        for col, week_start in sorted_weeks
        if week_start <= current_start_date
    ]

    if not current_start_candidates:
        return []

    first_current_week = current_start_candidates[-1][1]

    previous_weeks = [
        (col, week_start)
        for col, week_start in sorted_weeks
        if week_start < first_current_week
    ]

    selected = previous_weeks[-current_week_count:]

    return [col for col, week_start in selected]


def selected_columns_same_count_from_start(week_cols, target_start_date, target_week_count):
    """
    P2 rule.

    P2 must use the exact same number of counted weekly buckets
    as the Current Period.

    It maps the P2 start date to the latest week start on or before
    the P2 start date, then takes exactly N consecutive weekly buckets.
    """

    target_start_date = pd.to_datetime(target_start_date).normalize()

    sorted_weeks = sorted(
        [(col, pd.to_datetime(week_start).normalize()) for col, week_start in week_cols.items()],
        key=lambda x: x[1]
    )

    start_candidates = [
        (col, week_start)
        for col, week_start in sorted_weeks
        if week_start <= target_start_date
    ]

    if not start_candidates:
        return []

    mapped_start_week = start_candidates[-1][1]

    selected = [
        (col, week_start)
        for col, week_start in sorted_weeks
        if week_start >= mapped_start_week
    ][:target_week_count]

    return [col for col, week_start in selected]


def find_active_segment_for_week(week_start, segments):
    """
    Find which rollout segment should apply to a weekly bucket.

    Each weekly column represents a full 7-day bucket:
    week_start through week_start + 6 days.

    If a week overlaps multiple rollout phases, choose the phase
    with the highest number of overlapping days.

    Important:
    If a weekly bucket is selected as part of the valid period range
    but starts just before the calendar segment start date, still use
    the first applicable segment. This is required because weekly sales
    buckets can start before the exact P1/P2 calendar date.
    """

    week_start = pd.to_datetime(week_start).normalize()
    week_end = week_start + timedelta(days=6)

    best_segment = None
    best_overlap_days = 0

    for segment in segments:
        segment_start = pd.to_datetime(segment["start"]).normalize()
        segment_end = pd.to_datetime(segment["end"]).normalize()

        overlap_start = max(week_start, segment_start)
        overlap_end = min(week_end, segment_end)

        if overlap_start <= overlap_end:
            overlap_days = (overlap_end - overlap_start).days + 1

            if overlap_days > best_overlap_days:
                best_overlap_days = overlap_days
                best_segment = segment

    if best_segment is not None:
        return best_segment

    # Fallback:
    # If the selected weekly bucket is before the first segment,
    # apply the first segment's store rule.
    # This handles cases like P1 week 11/23/25 being selected
    # even though the P1 calendar range starts 12/1/25.
    if segments:
        first_segment = min(
            segments,
            key=lambda segment: pd.to_datetime(segment["start"]).normalize()
        )

        first_start = pd.to_datetime(first_segment["start"]).normalize()

        if week_start < first_start:
            return first_segment

    return None


def sum_section_for_segments(
    sheet,
    section_name,
    segments,
    period_name=None,
    current_week_count=None,
    current_start_date=None,
    p2_start_date=None
):
    """
    Mandatory week-count rule:

    Current determines the counted week count.

    P1 must use the exact same number of weeks as Current,
    immediately before the first Current week.

    P2 must also use the exact same number of weeks as Current,
    starting from the mapped P2 start week.

    Store coverage is applied week-by-week using rollout rules.
    """

    section_row, _ = find_row_containing(sheet, section_name)
    store_row, store_col = find_store_id_row(sheet, section_row)
    week_cols = get_week_columns(sheet, store_row)

    data_start = store_row + 1
    data_end = get_data_end_row(sheet, data_start, store_col)

    if period_name is None:
        period_name = segments[0]["period"] if segments else ""

    full_start = min(pd.to_datetime(segment["start"]).normalize() for segment in segments)
    full_end = max(pd.to_datetime(segment["end"]).normalize() for segment in segments)

    if period_name == "P1":
        selected_cols = selected_columns_before_current(
            week_cols,
            current_start_date,
            current_week_count
        )
    elif period_name == "P2":
        selected_cols = selected_columns_same_count_from_start(
            week_cols,
            p2_start_date,
            current_week_count
        )
    else:
        selected_cols = selected_columns_for_period(
            week_cols,
            full_start,
            full_end
        )

    selected_week_dates = [excel_date(week_cols[col]) for col in selected_cols]

    grand_total = 0
    store_names_used = set()
    debug_store_totals = {}
    week_debug_rows = []

    for col in selected_cols:
        week_start = week_cols[col]
        active_segment = find_active_segment_for_week(week_start, segments)

        if active_segment is None:
            continue

        allowed_stores = active_segment["stores"]

        week_total = 0
        week_store_count = 0

        for row in range(data_start, data_end):
            store_value = sheet.iat[row, store_col]

            if store_is_allowed(store_value, allowed_stores):
                numeric_value = pd.to_numeric(sheet.iat[row, col], errors="coerce")

                if pd.isna(numeric_value):
                    numeric_value = 0

                numeric_value = float(numeric_value)

                week_total += numeric_value
                grand_total += numeric_value

                store_name = clean_text(store_value)
                store_names_used.add(store_name)

                if store_name not in debug_store_totals:
                    debug_store_totals[store_name] = 0

                debug_store_totals[store_name] += numeric_value
                week_store_count += 1

        week_debug_rows.append({
            "Section": section_name,
            "Period": period_name,
            "Phase": active_segment["phase"],
            "Week Start": excel_date(week_start),
            "Stores Rule": "All Stores" if allowed_stores == "ALL" else ", ".join(allowed_stores),
            "Store Rows Used": week_store_count,
            "Week Total": week_total
        })

    debug_rows = []

    for store_name, store_total in debug_store_totals.items():
        debug_rows.append({
            "Section": section_name,
            "Period": period_name,
            "Phase": "Week-level rollout applied",
            "Segment Date Range": f"{excel_date(full_start)} - {excel_date(full_end)}",
            "Store Used": store_name,
            "Weeks Selected": ", ".join(selected_week_dates),
            "Week Count": len(selected_cols),
            "Store Total": store_total
        })

    segment_details = []

    for segment in segments:
        segment_details.append({
            "Period": segment["period"],
            "Phase": segment["phase"],
            "Date Range": f"{excel_date(segment['start'])} - {excel_date(segment['end'])}",
            "Stores": "All Stores" if segment["stores"] == "ALL" else ", ".join(segment["stores"]),
            "Store Rows Used": "",
            "Weeks Found": len(selected_cols),
            "Subtotal": ""
        })

    return {
        "total": float(grand_total),
        "weeks_found": len(selected_cols),
        "store_count": len(store_names_used),
        "store_names_used": sorted(store_names_used),
        "segment_details": segment_details,
        "debug_rows": debug_rows,
        "week_debug_rows": week_debug_rows
    }


def calculate_periods(start_date, end_date):
    start_date = pd.to_datetime(start_date).normalize()
    end_date = pd.to_datetime(end_date).normalize()

    duration_days = (end_date - start_date).days + 1

    p1_end = start_date - timedelta(days=1)
    p1_start = p1_end - timedelta(days=duration_days - 1)

    p2_start = start_date - pd.DateOffset(years=1)
    p2_end = end_date - pd.DateOffset(years=1)

    return {
        "Current": {"start": start_date, "end": end_date},
        "P1": {"start": p1_start, "end": p1_end},
        "P2": {"start": p2_start, "end": p2_end}
    }


def create_excel_report(campaign, sales_cpg, periods, unit_results, shopper_results, all_coverage_segments):
    wb = Workbook()

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = "Redemption"
    ws2 = wb.create_sheet("Pre-Post Analysis")
    ws3 = wb.create_sheet("Store Coverage Used")

    def style_range(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1["A1"] = "Redemption Summary"
    ws1["A1"].font = white_font
    ws1["A1"].fill = title_fill
    ws1.merge_cells("A1:B1")

    red_rows = [
        ("Offer Name", campaign.get("Offer Name", "")),
        ("Offer Period", f"{excel_date(campaign['Start Date'])} - {excel_date(campaign['End Date'])}"),
        ("Store Count Used", unit_results["Current"]["store_count"]),
        ("Offer Type", campaign.get("Type", "")),
        ("Activations", campaign.get("Activations", 0)),
        ("Redemptions", campaign.get("Redemptions", 0)),
        ("Redemption Ratio", "=IFERROR(B8/B7,0)"),
        ("CB Value", campaign.get("CB Value", 0)),
        ("Budget", campaign.get("Budget", 0)),
        ("Available Budget", campaign.get("Avail Budget", 0)),
        ("CB Amount", campaign.get("CB Amount", 0)),
        ("Unique Shoppers", campaign.get("Unique", 0)),
        ("New Shoppers", campaign.get("New", 0)),
    ]

    start_row = 3
    for idx, (label, value) in enumerate(red_rows, start=start_row):
        ws1[f"A{idx}"] = label
        ws1[f"B{idx}"] = value
        ws1[f"A{idx}"].font = bold_font
        ws1[f"A{idx}"].fill = header_fill

    ws1["B9"].number_format = "0%"

    for col in range(1, 3):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    style_range(ws1, f"A3:B{start_row + len(red_rows) - 1}")

    ws2["A1"] = f"Pre-Post Analysis - {sales_cpg}"
    ws2["A1"].font = white_font
    ws2["A1"].fill = title_fill
    ws2.merge_cells("A1:E1")

    weeks = unit_results["Current"]["weeks_found"]

    current_range = f"{excel_date(periods['Current']['start'])} - {excel_date(periods['Current']['end'])}"
    p1_range = f"{excel_date(periods['P1']['start'])} - {excel_date(periods['P1']['end'])}"
    p2_range = f"{excel_date(periods['P2']['start'])} - {excel_date(periods['P2']['end'])}"

    ws2["A3"] = "Unit Sales Calculation"
    ws2["A3"].font = white_font
    ws2["A3"].fill = title_fill
    ws2.merge_cells("A3:E3")

    headers = ["Previous Period", "Same Period Last Year", "Current Period", "Current vs P1", "Current vs P2"]

    for col, header in enumerate(headers, start=1):
        ws2.cell(row=5, column=col).value = header
        ws2.cell(row=5, column=col).font = bold_font
        ws2.cell(row=5, column=col).fill = header_fill

    ws2["A6"] = f"Previous {unit_results['P1']['weeks_found']} Weeks"
    ws2["B6"] = f"Same {unit_results['P2']['weeks_found']} Weeks Last Year"
    ws2["C6"] = f"Current {unit_results['Current']['weeks_found']} Weeks"

    ws2["A7"] = p1_range
    ws2["B7"] = p2_range
    ws2["C7"] = current_range

    ws2["A9"] = unit_results["P1"]["total"]
    ws2["B9"] = unit_results["P2"]["total"]
    ws2["C9"] = unit_results["Current"]["total"]
    ws2["D9"] = "=IFERROR((C9-A9)/A9,0)"
    ws2["E9"] = "=IFERROR((C9-B9)/B9,0)"

    ws2["D9"].number_format = "0%"
    ws2["E9"].number_format = "0%"

    style_range(ws2, "A5:E9")

    ws2["A12"] = "Unique Shopper Calculation"
    ws2["A12"].font = white_font
    ws2["A12"].fill = title_fill
    ws2.merge_cells("A12:E12")

    for col, header in enumerate(headers, start=1):
        ws2.cell(row=14, column=col).value = header
        ws2.cell(row=14, column=col).font = bold_font
        ws2.cell(row=14, column=col).fill = header_fill

    ws2["A15"] = f"Previous {shopper_results['P1']['weeks_found']} Weeks"
    ws2["B15"] = f"Same {shopper_results['P2']['weeks_found']} Weeks Last Year"
    ws2["C15"] = f"Current {shopper_results['Current']['weeks_found']} Weeks"

    ws2["A16"] = p1_range
    ws2["B16"] = p2_range
    ws2["C16"] = current_range

    ws2["A18"] = shopper_results["P1"]["total"]
    ws2["B18"] = shopper_results["P2"]["total"]
    ws2["C18"] = shopper_results["Current"]["total"]
    ws2["D18"] = "=IFERROR((C18-A18)/A18,0)"
    ws2["E18"] = "=IFERROR((C18-B18)/B18,0)"

    ws2["D18"].number_format = "0%"
    ws2["E18"].number_format = "0%"

    style_range(ws2, "A14:E18")

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 26

    ws3["A1"] = "Store Coverage Used"
    ws3["A1"].font = white_font
    ws3["A1"].fill = title_fill
    ws3.merge_cells("A1:G1")

    coverage_headers = ["Period", "Phase", "Date Range", "Stores", "Store Rows Used", "Weeks Found", "Subtotal Rule"]
    for col, header in enumerate(coverage_headers, start=1):
        ws3.cell(row=3, column=col).value = header
        ws3.cell(row=3, column=col).font = bold_font
        ws3.cell(row=3, column=col).fill = header_fill

    row_num = 4
    for segment in all_coverage_segments:
        ws3.cell(row=row_num, column=1).value = segment["period"]
        ws3.cell(row=row_num, column=2).value = segment["phase"]
        ws3.cell(row=row_num, column=3).value = f"{excel_date(segment['start'])} - {excel_date(segment['end'])}"
        ws3.cell(row=row_num, column=4).value = "All Stores" if segment["stores"] == "ALL" else ", ".join(segment["stores"])
        ws3.cell(row=row_num, column=5).value = ""
        ws3.cell(row=row_num, column=6).value = ""
        ws3.cell(row=row_num, column=7).value = "Coverage determined by the period being analyzed"
        row_num += 1

    style_range(ws3, f"A3:G{row_num - 1}")

    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
from rollout_rules import (
    build_current_coverage_segments,
    build_p1_coverage_segments,
    build_p2_coverage_segments,
    store_is_allowed,
)

st.title("📊 Retail Campaign Intelligence Engine")
st.caption("Offline campaign analysis | P1 weekly bucket fallback enabled")

st.divider()

left, right = st.columns(2)

with left:
    st.subheader("1️⃣ Campaign Library")
    redemption_file = st.file_uploader(
        "Upload Daily Redemption Report only when new redemption data is added",
        type=["xlsx"]
    )

    if redemption_file:
        try:
            lib = load_redemption_report(redemption_file)
            st.success(f"✅ Campaign Library Saved: {len(lib)} campaign(s)")
            st.dataframe(lib[["CPG", "Offer Name", "Start Date", "End Date"]].head(10), use_container_width=True)
        except Exception as e:
            st.error(f"Redemption Error: {e}")

    if os.path.exists(LIBRARY_FILE):
        st.info("✅ Campaign Library is available. You can now upload Sales Report only.")

with right:
    st.subheader("2️⃣ Sales Report")
    sales_file = st.file_uploader(
        "Upload Sales Report for analysis",
        type=["xlsx"]
    )

st.divider()
st.divider()

st.subheader("🚀 Batch Analysis")

batch_mode = st.checkbox("Enable Batch Analysis")

batch_folder = st.text_input(
    "Sales Report Folder",
    value=r"C:\Users\ADMIN\Kacu Technology Inc\Partnership Development - Documents\Attachments\CPG Sales Numbers"
)

batch_limit = st.number_input(
    "Maximum CPGs",
    min_value=1,
    max_value=500,
    value=25
)

batch_working_folder = st.text_input(
    "Local Batch Working Folder",
    value=r"C:\Users\ADMIN\Desktop\Batch Working"
)

prepare_batch = st.button("📂 Prepare Batch Folder")

run_batch = st.button("🚀 Run Batch Analysis")

from pathlib import Path
import shutil

if prepare_batch:
    try:
        source_folder = Path(batch_folder)
        target_folder = Path(batch_working_folder)

        if not source_folder.exists():
            st.error("Source Sales Report folder not found.")
            st.stop()

        target_folder.mkdir(parents=True, exist_ok=True)

        library_df = load_campaign_library()

        sales_files = [
            file for file in source_folder.glob("*.xlsx")
            if not file.name.startswith("~$")
        ]

        copied = []
        missing = []
        errors = []

        for _, campaign in library_df.iterrows():
            cpg_name = str(campaign.get("CPG", "")).strip()
            target_cpg = normalize_cpg_name(cpg_name)

            matched_file = None

            for file in sales_files:
                file_name_clean = normalize_cpg_name(file.stem)

                if file_name_clean == target_cpg:
                    matched_file = file
                    break

            if matched_file is None:
                for file in sales_files:
                    file_name_clean = normalize_cpg_name(file.stem)

                    if target_cpg in file_name_clean or file_name_clean in target_cpg:
                        matched_file = file
                        break

            if matched_file is None:
                missing.append({
                    "CPG": cpg_name,
                    "Reason": "No matching Sales Report filename found"
                })
                continue

            try:
                destination = target_folder / matched_file.name
                shutil.copy2(str(matched_file), str(destination))

                copied.append({
                    "CPG": cpg_name,
                    "Copied File": matched_file.name
                })

            except Exception as e:
                errors.append({
                    "CPG": cpg_name,
                    "Matched File": matched_file.name,
                    "Error": str(e)
                })

        st.success(f"Prepare complete. Copied: {len(copied)}, Missing: {len(missing)}, Errors: {len(errors)}")
        st.info(f"Prepared folder: {target_folder}")

        if copied:
            st.subheader("✅ Copied")
            st.dataframe(pd.DataFrame(copied), use_container_width=True)

        if missing:
            st.subheader("⚠️ Missing")
            st.dataframe(pd.DataFrame(missing), use_container_width=True)

        if errors:
            st.subheader("❌ Copy Errors")
            st.dataframe(pd.DataFrame(errors), use_container_width=True)

    except Exception as e:
        st.error(f"Prepare Batch Folder Error: {e}")


def find_sales_file_for_cpg_batch(cpg_name, folder):
    target = normalize_cpg_name(cpg_name)
    files = [f for f in Path(folder).glob("*.xlsx") if not f.name.startswith("~$")]

    for f in files:
        name = normalize_cpg_name(f.stem)
        if name == target:
            return f

    for f in files:
        name = normalize_cpg_name(f.stem)
        if target in name or name in target:
            return f

    return None


def save_batch_result_to_master(campaign, sales_cpg, unit_results, shopper_results, unit_vs_p1, unit_vs_p2, shopper_vs_p1, shopper_vs_p2):
    from openpyxl import load_workbook

    template_path = Path("Output Format.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    cpg_name_value = str(campaign.get("CPG", sales_cpg)).strip()
    offer_name_value = str(campaign.get("Offer Name", "")).strip()
    offer_start_value = pd.to_datetime(campaign["Start Date"])
    offer_end_value = pd.to_datetime(campaign["End Date"])

    row = 2
    while ws.cell(row, 1).value not in [None, ""]:
        row += 1

    redemption_ratio = 0
    try:
        if campaign.get("Activations", 0) not in [0, "", None]:
            redemption_ratio = campaign.get("Redemptions", 0) / campaign.get("Activations", 0)
    except Exception:
        redemption_ratio = 0

    try:
        budget_consumed = campaign.get("Budget", 0) - campaign.get("Avail Budget", 0)
    except Exception:
        budget_consumed = ""

    row_values = [
        cpg_name_value,
        offer_name_value,
        campaign.get("Type", ""),
        unit_results["Current"]["store_count"],
        offer_start_value,
        offer_end_value,
        campaign.get("Activations", 0),
        campaign.get("Redemptions", 0),
        redemption_ratio,
        campaign.get("CB Value", 0),
        campaign.get("Budget", 0),
        campaign.get("Avail Budget", 0),
        budget_consumed,
        campaign.get("Unique", 0),
        campaign.get("New", 0),
        unit_results["P1"]["total"],
        unit_results["P2"]["total"],
        unit_results["Current"]["total"],
        unit_vs_p1 / 100 if unit_vs_p1 is not None else "",
        unit_vs_p2 / 100 if unit_vs_p2 is not None else "",
        shopper_results["P1"]["total"],
        shopper_results["P2"]["total"],
        shopper_results["Current"]["total"],
        shopper_vs_p1 / 100 if shopper_vs_p1 is not None else "",
        shopper_vs_p2 / 100 if shopper_vs_p2 is not None else "",
    ]

    for col, value in enumerate(row_values, start=1):
        ws.cell(row, col).value = "" if pd.isna(value) else value

    ws.cell(row, 5).number_format = "m/d/yy"
    ws.cell(row, 6).number_format = "m/d/yy"

    for pct_col in [9, 19, 20, 24, 25]:
        ws.cell(row, pct_col).number_format = "0.0%"

    wb.save(template_path)
    return row


if run_batch:
    try:
        folder = Path(batch_folder)

        if not folder.exists():
            st.error("Sales Report folder not found.")
            st.stop()

        if not Path("Output Format.xlsx").exists():
            st.error("Output Format.xlsx not found. Please keep it in the project folder.")
            st.stop()

        library_df = load_campaign_library()
        campaigns_to_process = library_df.head(int(batch_limit))

        processed = []
        missing = []
        errors = []

        progress = st.progress(0)
        status = st.empty()

        total = len(campaigns_to_process)

        for idx, (_, campaign) in enumerate(campaigns_to_process.iterrows(), start=1):
            cpg_name = str(campaign.get("CPG", "")).strip()
            status.write(f"Processing {idx}/{total}: {cpg_name}")

            sales_file = find_sales_file_for_cpg_batch(cpg_name, folder)

            if sales_file is None:
                missing.append({"CPG": cpg_name, "Reason": "Sales report not found"})
                progress.progress(idx / total)
                continue

            try:
                sales_sheet = read_sales_sheet(str(sales_file))
                sales_cpg = get_sales_cpg(sales_sheet)

                periods = calculate_periods(campaign["Start Date"], campaign["End Date"])

                period_coverage_segments = {
                    "Current": build_current_coverage_segments(periods["Current"]["start"], periods["Current"]["end"]),
                    "P1": build_p1_coverage_segments(periods["P1"]["start"], periods["P1"]["end"]),
                    "P2": build_p2_coverage_segments(periods["P2"]["start"], periods["P2"]["end"]),
                }

                unit_results = {}
                shopper_results = {}

                unit_results["Current"] = sum_section_for_segments(
                    sales_sheet, "Unit Sales by Store", period_coverage_segments["Current"], period_name="Current"
                )

                shopper_results["Current"] = sum_section_for_segments(
                    sales_sheet, "Unique Shoppers by Store", period_coverage_segments["Current"], period_name="Current"
                )

                current_week_count = unit_results["Current"]["weeks_found"]
                current_start_date = periods["Current"]["start"]
                p2_start_date = periods["P2"]["start"]

                unit_results["P1"] = sum_section_for_segments(
                    sales_sheet, "Unit Sales by Store", period_coverage_segments["P1"],
                    period_name="P1", current_week_count=current_week_count, current_start_date=current_start_date
                )

                shopper_results["P1"] = sum_section_for_segments(
                    sales_sheet, "Unique Shoppers by Store", period_coverage_segments["P1"],
                    period_name="P1", current_week_count=current_week_count, current_start_date=current_start_date
                )

                unit_results["P2"] = sum_section_for_segments(
                    sales_sheet, "Unit Sales by Store", period_coverage_segments["P2"],
                    period_name="P2", current_week_count=current_week_count, p2_start_date=p2_start_date
                )

                shopper_results["P2"] = sum_section_for_segments(
                    sales_sheet, "Unique Shoppers by Store", period_coverage_segments["P2"],
                    period_name="P2", current_week_count=current_week_count, p2_start_date=p2_start_date
                )

                unit_vs_p1 = pct_change(unit_results["Current"]["total"], unit_results["P1"]["total"])
                unit_vs_p2 = pct_change(unit_results["Current"]["total"], unit_results["P2"]["total"])
                shopper_vs_p1 = pct_change(shopper_results["Current"]["total"], shopper_results["P1"]["total"])
                shopper_vs_p2 = pct_change(shopper_results["Current"]["total"], shopper_results["P2"]["total"])

                row_number = save_batch_result_to_master(
                    campaign, sales_cpg, unit_results, shopper_results,
                    unit_vs_p1, unit_vs_p2, shopper_vs_p1, shopper_vs_p2
                )

                processed.append({
                    "CPG": cpg_name,
                    "Sales File": sales_file.name,
                    "Row": row_number,
                    "Current Units": unit_results["Current"]["total"],
                    "Current Shoppers": shopper_results["Current"]["total"],
                })

            except Exception as e:
                errors.append({"CPG": cpg_name, "Sales File": sales_file.name, "Error": str(e)})

            progress.progress(idx / total)

        st.success(f"Batch completed. Processed: {len(processed)}, Missing: {len(missing)}, Errors: {len(errors)}")

        if processed:
            st.subheader("✅ Processed")
            st.dataframe(pd.DataFrame(processed), use_container_width=True)

        if missing:
            st.subheader("⚠️ Missing")
            st.dataframe(pd.DataFrame(missing), use_container_width=True)

        if errors:
            st.subheader("❌ Errors")
            st.dataframe(pd.DataFrame(errors), use_container_width=True)

        with open("Output Format.xlsx", "rb") as file:
            st.download_button(
                label="📥 Download Updated Master Workbook",
                data=file,
                file_name="Output Format.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    except Exception as e:
        st.error(f"Batch Analysis Error: {e}")


if sales_file:
    if st.button("🚀 Analyze Campaign", use_container_width=True):
        try:
            library_df = load_campaign_library()
            sales_sheet = read_sales_sheet(sales_file)
            sales_cpg = get_sales_cpg(sales_sheet)
            campaign, match_type = find_campaign(library_df, sales_cpg)

            periods = calculate_periods(campaign["Start Date"], campaign["End Date"])

            period_coverage_segments = {
                "Current": build_current_coverage_segments(periods["Current"]["start"], periods["Current"]["end"]),
                "P1": build_p1_coverage_segments(periods["P1"]["start"], periods["P1"]["end"]),
                "P2": build_p2_coverage_segments(periods["P2"]["start"], periods["P2"]["end"]),
            }

            unit_results = {}
            shopper_results = {}

            # Current is calculated first.
            # Its week count becomes the mandatory week count for P1 and P2.
            unit_results["Current"] = sum_section_for_segments(
                sales_sheet,
                "Unit Sales by Store",
                period_coverage_segments["Current"],
                period_name="Current"
            )

            shopper_results["Current"] = sum_section_for_segments(
                sales_sheet,
                "Unique Shoppers by Store",
                period_coverage_segments["Current"],
                period_name="Current"
            )

            current_week_count = unit_results["Current"]["weeks_found"]
            current_start_date = periods["Current"]["start"]
            p2_start_date = periods["P2"]["start"]

            # P1 must use the same week count as Current,
            # immediately before the first Current week.
            unit_results["P1"] = sum_section_for_segments(
                sales_sheet,
                "Unit Sales by Store",
                period_coverage_segments["P1"],
                period_name="P1",
                current_week_count=current_week_count,
                current_start_date=current_start_date
            )

            shopper_results["P1"] = sum_section_for_segments(
                sales_sheet,
                "Unique Shoppers by Store",
                period_coverage_segments["P1"],
                period_name="P1",
                current_week_count=current_week_count,
                current_start_date=current_start_date
            )

            # P2 must also use the same week count as Current.
            unit_results["P2"] = sum_section_for_segments(
                sales_sheet,
                "Unit Sales by Store",
                period_coverage_segments["P2"],
                period_name="P2",
                current_week_count=current_week_count,
                p2_start_date=p2_start_date
            )

            shopper_results["P2"] = sum_section_for_segments(
                sales_sheet,
                "Unique Shoppers by Store",
                period_coverage_segments["P2"],
                period_name="P2",
                current_week_count=current_week_count,
                p2_start_date=p2_start_date
            )

            unit_vs_p1 = pct_change(unit_results["Current"]["total"], unit_results["P1"]["total"])
            unit_vs_p2 = pct_change(unit_results["Current"]["total"], unit_results["P2"]["total"])
            shopper_vs_p1 = pct_change(shopper_results["Current"]["total"], shopper_results["P1"]["total"])
            shopper_vs_p2 = pct_change(shopper_results["Current"]["total"], shopper_results["P2"]["total"])

            st.success("✅ Campaign Analysis Completed")

            st.subheader("📋 Campaign Identified")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Sales Report CPG", sales_cpg)
            c2.metric("Offer Type", campaign["Type"])
            c3.metric("Match Method", match_type)
            c4.metric("Current Stores Counted", unit_results["Current"]["store_count"])

            st.write(f"**Campaign Library CPG:** {campaign['CPG']}")
            st.write(f"**Offer Name:** {campaign['Offer Name']}")
            st.write(f"**Offer Period:** {excel_date(campaign['Start Date'])} - {excel_date(campaign['End Date'])}")

            st.subheader("🏬 Store Coverage Used")

            coverage_rows = []
            for period_name in ["Current", "P1", "P2"]:
                for segment in period_coverage_segments[period_name]:
                    coverage_rows.append({
                        "Period": period_name,
                        "Phase": segment["phase"],
                        "Date Range": f"{excel_date(segment['start'])} - {excel_date(segment['end'])}",
                        "Stores": "All Stores" if segment["stores"] == "ALL" else ", ".join(segment["stores"])
                    })

            st.dataframe(pd.DataFrame(coverage_rows), use_container_width=True)

            result_df = pd.DataFrame([
                {
                    "Metric": "Unit Sales",
                    "Current": unit_results["Current"]["total"],
                    "P1": unit_results["P1"]["total"],
                    "P2": unit_results["P2"]["total"],
                    "Current vs P1": format_pct(unit_vs_p1),
                    "Current vs P2": format_pct(unit_vs_p2),
                },
                {
                    "Metric": "Unique Shoppers",
                    "Current": shopper_results["Current"]["total"],
                    "P1": shopper_results["P1"]["total"],
                    "P2": shopper_results["P2"]["total"],
                    "Current vs P1": format_pct(shopper_vs_p1),
                    "Current vs P2": format_pct(shopper_vs_p2),
                }
            ])

            st.dataframe(result_df, use_container_width=True)

            from openpyxl import load_workbook
            from pathlib import Path

            template_path = Path("Output Format.xlsx")

            if template_path.exists():
                wb = load_workbook(template_path)
                ws = wb.active

                def safe_value(value):
                    if pd.isna(value):
                        return ""
                    return value

                def normalize_key_value(value):
                    if value is None:
                        return ""
                    try:
                        return pd.to_datetime(value).strftime("%Y-%m-%d")
                    except Exception:
                        return str(value).strip().lower()

                cpg_name_value = str(campaign.get("CPG", sales_cpg)).strip()
                offer_name_value = str(campaign.get("Offer Name", "")).strip()
                offer_start_value = pd.to_datetime(campaign["Start Date"])
                offer_end_value = pd.to_datetime(campaign["End Date"])

                # Unique key: CPG + Offer Name + Offer Start + Offer End
                target_row = None
                row = 2

                while ws.cell(row, 1).value not in [None, ""]:
                    existing_cpg = str(ws.cell(row, 1).value).strip().lower()
                    existing_offer = str(ws.cell(row, 2).value).strip().lower()
                    existing_start = normalize_key_value(ws.cell(row, 5).value)
                    existing_end = normalize_key_value(ws.cell(row, 6).value)

                    if (
                        existing_cpg == cpg_name_value.lower()
                        and existing_offer == offer_name_value.lower()
                        and existing_start == offer_start_value.strftime("%Y-%m-%d")
                        and existing_end == offer_end_value.strftime("%Y-%m-%d")
                    ):
                        target_row = row
                        break

                    row += 1

                if target_row is None:
                    target_row = row
                    action_text = "added"
                else:
                    action_text = "updated"

                redemption_ratio = 0
                if campaign.get("Activations", 0) not in [0, "", None]:
                    try:
                        redemption_ratio = campaign.get("Redemptions", 0) / campaign.get("Activations", 0)
                    except Exception:
                        redemption_ratio = 0

                budget_consumed = 0
                try:
                    budget_consumed = campaign.get("Budget", 0) - campaign.get("Avail Budget", 0)
                except Exception:
                    budget_consumed = ""

                row_values = [
                    cpg_name_value,
                    offer_name_value,
                    campaign.get("Type", ""),
                    unit_results["Current"]["store_count"],
                    offer_start_value,
                    offer_end_value,
                    campaign.get("Activations", 0),
                    campaign.get("Redemptions", 0),
                    redemption_ratio,
                    campaign.get("CB Value", 0),
                    campaign.get("Budget", 0),
                    campaign.get("Avail Budget", 0),
                    budget_consumed,
                    campaign.get("Unique", 0),
                    campaign.get("New", 0),
                    unit_results["P1"]["total"],
                    unit_results["P2"]["total"],
                    unit_results["Current"]["total"],
                    unit_vs_p1 / 100 if unit_vs_p1 is not None else "",
                    unit_vs_p2 / 100 if unit_vs_p2 is not None else "",
                    shopper_results["P1"]["total"],
                    shopper_results["P2"]["total"],
                    shopper_results["Current"]["total"],
                    shopper_vs_p1 / 100 if shopper_vs_p1 is not None else "",
                    shopper_vs_p2 / 100 if shopper_vs_p2 is not None else "",
                ]

                for col_idx, value in enumerate(row_values, start=1):
                    ws.cell(target_row, col_idx).value = safe_value(value)

                # Date formatting
                ws.cell(target_row, 5).number_format = "m/d/yy"
                ws.cell(target_row, 6).number_format = "m/d/yy"

                # Percentage formatting
                for pct_col in [9, 19, 20, 24, 25]:
                    ws.cell(target_row, pct_col).number_format = "0.0%"

                wb.save(template_path)

                st.success(f"✅ Master workbook {action_text}: row {target_row}")
                st.info("📄 Output Format.xlsx has been updated and saved permanently.")

            else:
                st.error("Output Format.xlsx not found in project folder.")


            st.subheader("🔍 P1 Debug Details")

            with st.expander("P1 Debug - Unit Sales", expanded=False):
                p1_unit_debug = pd.DataFrame(unit_results["P1"].get("debug_rows", []))
                p1_unit_week_debug = pd.DataFrame(unit_results["P1"].get("week_debug_rows", []))

                st.write("Store-level P1 Unit Debug")
                if not p1_unit_debug.empty:
                    st.dataframe(p1_unit_debug, use_container_width=True)
                    st.write("P1 Unit Sales Store Total:", p1_unit_debug["Store Total"].sum())
                else:
                    st.warning("No P1 Unit Sales store debug rows found.")

                st.write("Week-level P1 Unit Debug")
                if not p1_unit_week_debug.empty:
                    st.dataframe(p1_unit_week_debug, use_container_width=True)
                    st.write("P1 Unit Sales Week Total:", p1_unit_week_debug["Week Total"].sum())
                else:
                    st.warning("No P1 Unit Sales week debug rows found.")

            with st.expander("P1 Debug - Unique Shoppers", expanded=False):
                p1_shopper_debug = pd.DataFrame(shopper_results["P1"].get("debug_rows", []))
                p1_shopper_week_debug = pd.DataFrame(shopper_results["P1"].get("week_debug_rows", []))

                st.write("Store-level P1 Shopper Debug")
                if not p1_shopper_debug.empty:
                    st.dataframe(p1_shopper_debug, use_container_width=True)
                    st.write("P1 Unique Shoppers Store Total:", p1_shopper_debug["Store Total"].sum())
                else:
                    st.warning("No P1 Unique Shopper store debug rows found.")

                st.write("Week-level P1 Shopper Debug")
                if not p1_shopper_week_debug.empty:
                    st.dataframe(p1_shopper_week_debug, use_container_width=True)
                    st.write("P1 Unique Shoppers Week Total:", p1_shopper_week_debug["Week Total"].sum())
                else:
                    st.warning("No P1 Unique Shopper week debug rows found.")

            all_segments = []
            for period_name in ["Current", "P1", "P2"]:
                all_segments.extend(period_coverage_segments[period_name])

            template_path = Path("Output Format.xlsx")

            if template_path.exists():
                with open(template_path, "rb") as file:
                    st.download_button(
                        label="📥 Download Updated Master Analysis Workbook",
                        data=file,
                        file_name="Output Format.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.error("Output Format.xlsx not found for download.")

        except Exception as e:
            st.error(f"Analysis Error: {e}")
else:
    st.info("Upload the Sales Report to begin analysis.")